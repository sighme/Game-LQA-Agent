import json
import re
import sys
import openpyxl
from openpyxl.styles import Font, Alignment


def extract_eng_texts(filepath):
    """Extract all ENG field values from a JSON file using regex (handles malformed JSON too)."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    # Match "ENG": "some text" — captures content between the quotes after ENG:
    pattern = r'"ENG"\s*:\s*"((?:[^"\\]|\\.)*)"'
    matches = re.findall(pattern, content)
    # Unescape any escaped characters
    return [m.replace('\\"', '"').replace('\\\\', '\\') for m in matches]


def main():
    if len(sys.argv) < 3:
        print("Usage: python extract_subtitles.py <english_json> <chinese_json>")
        print("Example: python extract_subtitles.py en.json zh.json")
        sys.exit(1)

    en_file = sys.argv[1]
    zh_file = sys.argv[2]

    en_texts = extract_eng_texts(en_file)
    zh_texts = extract_eng_texts(zh_file)

    if len(en_texts) != len(zh_texts):
        print(f"Warning: English has {len(en_texts)} entries, Chinese has {len(zh_texts)} entries.")
        print("Rows will be paired up to the shorter list; extras will be left blank.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subtitles"

    header_font = Font(name='Arial', bold=True)
    ws['A1'] = 'Source (English)'
    ws['B1'] = 'Translation (Chinese)'
    ws['A1'].font = header_font
    ws['B1'].font = header_font

    cell_font = Font(name='Arial', size=11)
    wrap = Alignment(wrap_text=True, vertical='top')

    max_rows = max(len(en_texts), len(zh_texts))
    for i in range(max_rows):
        row = i + 2
        en_val = en_texts[i] if i < len(en_texts) else ''
        zh_val = zh_texts[i] if i < len(zh_texts) else ''
        ws.cell(row=row, column=1, value=en_val).font = cell_font
        ws.cell(row=row, column=2, value=zh_val).font = cell_font
        ws.cell(row=row, column=1).alignment = wrap
        ws.cell(row=row, column=2).alignment = wrap

    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 60

    output = 'subtitles_bilingual.xlsx'
    wb.save(output)
    print(f"Done! {max_rows} rows written to {output}")
    print(f"  English entries: {len(en_texts)}")
    print(f"  Chinese entries: {len(zh_texts)}")


if __name__ == '__main__':
    main()